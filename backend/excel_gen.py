import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_excel(data: List[Dict[str, Any]], metadata_info: Dict[str, Any]) -> io.BytesIO:
    wb = Workbook()
    
    # 1. Sheet 'Data'
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Stylings
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid") # Classic Excel Blue
    center_align = Alignment(horizontal="center", vertical="center")
    default_font = Font(name="Calibri", size=11)
    
    # Write Headers
    if data:
        headers = list(data[0].keys())
    else:
        headers = metadata_info.get("field_names", ["No Data"])
        
    ws_data.append(headers)
    
    # Style Headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Write Rows
    for row in data:
        row_vals = [row.get(h, "") for h in headers]
        ws_data.append(row_vals)
        
    # Apply standard fonts
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=1, max_col=ws_data.max_column):
        for cell in row:
            cell.font = default_font

    # Auto-fit columns for Data sheet
    for col in ws_data.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. Sheet 'Metadata'
    ws_meta = wb.create_sheet(title="Metadata")
    
    # Headers for metadata
    ws_meta.append(["Property", "Value"])
    for col_idx in range(1, 3):
        cell = ws_meta.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    meta_rows = [
        ["Job ID", metadata_info.get("job_id", "N/A")],
        ["Job Name", metadata_info.get("job_name", "N/A")],
        ["Target URL", metadata_info.get("url", "N/A")],
        ["Date Scraped", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Rows Scraped", len(data)],
        ["Max Pages Config", metadata_info.get("max_pages", "N/A")],
        ["Pagination Mode", metadata_info.get("pagination_type", "N/A")]
    ]
    
    for row in meta_rows:
        ws_meta.append(row)
        
    for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row, min_col=1, max_col=ws_meta.max_column):
        for cell in row:
            cell.font = default_font

    # Auto-fit columns for Metadata sheet
    for col in ws_meta.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws_meta.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Save workbook to memory stream
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream
